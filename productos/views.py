# ── Importaciones necesarias ─────────────────────────────────────────────────
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required      # Protege vistas de funciones
from django.contrib.auth.mixins import LoginRequiredMixin      # Protege vistas de clases
from django.views.generic import ListView
from fpdf import FPDF
from openpyxl import Workbook
from .models import Producto
from .forms import ProductoForm

# ── LISTADO ───────────────────────────────────────────────────────────────────
class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = 'productos/listar_producto.html'
    context_object_name = 'productos'  # nombre de la variable en el template
    paginate_by = 4                   # 4 productos por página
    ordering = ['productoId']


# ── CREAR ─────────────────────────────────────────────────────────────────────
@login_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()                # Guarda el nuevo producto en la BD
            return redirect('listar_productos')
    else:
        form = ProductoForm()          # Formulario vacío
    return render(request, 'productos/crear_producto.html', {'form': form})


# ── VER DETALLE ───────────────────────────────────────────────────────────────
@login_required
def ver_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'productos/ver_producto.html', {'producto': producto})


# ── ACTUALIZAR ────────────────────────────────────────────────────────────────
@login_required
def actualizar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)  # Vincula el form al registro
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)  # Muestra datos actuales del producto
    return render(request, 'productos/actualizar_producto.html', {'form': form, 'producto': producto})


# ── ELIMINAR ──────────────────────────────────────────────────────────────────
# Se pide confirmación vía POST antes de eliminar (el GET solo muestra el template de confirmación).
@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    error = None
    if request.method == 'POST':       # Solo elimina si el usuario confirmó
        try:
            producto.delete()
            return redirect('listar_productos')
        except Exception as e:
            error = 'No se puede eliminar el producto porque está asociado a uno o más pedidos.'
    return render(request, 'productos/eliminar_producto.html', {'producto': producto, 'error': error})
    return render(request, 'productos/eliminar_producto.html', {'producto': producto})


@login_required
def exportar_productos_pdf(request):
    def limpiar(texto):
        reemplazos = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n',
                      'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ñ':'N'}
        for k, v in reemplazos.items():
            texto = texto.replace(k, v)
        return texto

    productos = Producto.objects.all()
    pdf = FPDF()
    pdf.add_page()
    ancho = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(ancho, 10, 'Reporte de Productos', ln=True)
    # Cabecera de tabla
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(15, 8, 'ID', 1)
    pdf.cell(60, 8, 'Nombre', 1)
    pdf.cell(30, 8, 'Precio', 1)
    pdf.cell(25, 8, 'Stock', 1)
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    for producto in productos:
        pdf.cell(15, 8, str(producto.productoId), 1)
        pdf.cell(60, 8, limpiar(producto.nombre_producto), 1)
        pdf.cell(30, 8, f"${producto.precio:.2f}", 1)
        pdf.cell(25, 8, str(producto.unidades_stock), 1)
        pdf.ln()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
    response.write(bytes(pdf.output(dest='S')))
    return response


@login_required
def exportar_productos_excel(request):
    productos = Producto.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    ws.append(['ID', 'Nombre', 'Precio', 'Stock'])
    for producto in productos:
        ws.append([
            producto.productoId,
            producto.nombre_producto,
            float(producto.precio),
            producto.unidades_stock,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response
