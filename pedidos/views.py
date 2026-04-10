# ── Importaciones necesarias ─────────────────────────────────────────────────
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.forms import inlineformset_factory   # Permite manejar múltiples forms relacionados
from django.http import HttpResponse
from django.views.generic import ListView
from django.shortcuts import render, get_object_or_404, redirect
from fpdf import FPDF
from openpyxl import Workbook
from .models import Pedido
from .forms import PedidoForm
from detalles_pedido.models import DetallePedido

# ── FORMSET INLINE ────────────────────────────────────────────────────────────
# Permite agregar múltiples productos a un pedido desde el mismo formulario.
# extra=1 muestra 1 fila vacía por defecto; can_delete=True agrega checkbox para eliminar filas.
DetalleInlineFormSet = inlineformset_factory(
    Pedido,           # Modelo padre
    DetallePedido,    # Modelo hijo
    fields=['productoId', 'cantidad'],
    fk_name='pedidoId',
    extra=1,
    can_delete=True,
)

# ── LISTADO ───────────────────────────────────────────────────────────────────
# select_related('clienteId') hace un JOIN con la tabla Cliente en una sola consulta SQL
# en lugar de hacer una consulta extra por cada pedido (optimización N+1).
class PedidoListView(LoginRequiredMixin, ListView):
    template_name = 'pedidos/listar_pedidos.html'
    context_object_name = 'pedidos'
    paginate_by = 4

    def get_queryset(self):
        return Pedido.objects.select_related('clienteId').all().order_by('pedidoId')


# ── CREAR ─────────────────────────────────────────────────────────────────────
# Se validan tanto el form del pedido como el formset de detalles antes de guardar.
@login_required
def crear_pedido(request):
    error = None
    if request.method == 'POST':
        form    = PedidoForm(request.POST)
        formset = DetalleInlineFormSet(request.POST, prefix='detalles')
        if form.is_valid() and formset.is_valid():
            try:
                pedido           = form.save()       # Guarda el pedido primero
                formset.instance = pedido            # Vincula los detalles al pedido recién creado
                formset.save()                       # Guarda los detalles en la BD
                return redirect('listar_pedidos')
            except Exception as e:
                error = str(e)
        # Si hay errores de validación en el formset, los errores ya estarán en los forms
    else:
        form    = PedidoForm()
        formset = DetalleInlineFormSet(prefix='detalles')  # Formset vacío
    return render(request, 'pedidos/crear_pedido.html', {'form': form, 'formset': formset, 'error': error})


# ── VER DETALLE ───────────────────────────────────────────────────────────────
@login_required
def ver_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    return render(request, 'pedidos/ver_pedido.html', {'pedido': pedido})


# ── ACTUALIZAR ────────────────────────────────────────────────────────────────
# instance=pedido vincula el form y el formset al pedido existente para editarlo.
@login_required
def actualizar_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    if request.method == 'POST':
        form    = PedidoForm(request.POST, instance=pedido)
        formset = DetalleInlineFormSet(request.POST, instance=pedido, prefix='detalles')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()  # Actualiza, crea o elimina detalles según lo marcado
            return redirect('listar_pedidos')
    else:
        form    = PedidoForm(instance=pedido)
        formset = DetalleInlineFormSet(instance=pedido, prefix='detalles')
    return render(request, 'pedidos/actualizar_pedido.html', {'form': form, 'pedido': pedido, 'formset': formset})


# ── ELIMINAR ──────────────────────────────────────────────────────────────────
# Regla de negocio: no se puede eliminar un pedido que tenga detalles asociados.
# Se debe eliminar primero cada DetallePedido (para que el stock se restaure correctamente).
@login_required
def eliminar_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    if pedido.detalles.exists():  # Verifica si hay detalles relacionados
        from django.contrib import messages
        messages.error(request, 'No se puede eliminar el pedido porque tiene productos asociados. Elimine primero los detalles.')
        return redirect('listar_pedidos')
    if request.method == 'POST':
        pedido.delete()
        return redirect('listar_pedidos')
    return render(request, 'pedidos/eliminar_pedido.html', {'pedido': pedido})


@login_required
def exportar_pedidos_pdf(request):
    def limpiar(texto):
        reemplazos = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n',
                      'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ñ':'N'}
        for k, v in reemplazos.items():
            texto = texto.replace(k, v)
        return texto

    pedidos = Pedido.objects.select_related('clienteId').all()
    pdf = FPDF()
    pdf.add_page()
    ancho = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(ancho, 10, 'Reporte de Pedidos', ln=True)
    pdf.set_font('Arial', '', 10)
    # Cabecera de tabla
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(20, 8, 'ID', 1)
    pdf.cell(50, 8, 'Cliente', 1)
    pdf.cell(35, 8, 'Fecha', 1)
    pdf.cell(35, 8, 'Estado', 1)
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    for pedido in pedidos:
        pdf.cell(20, 8, str(pedido.pedidoId), 1)
        pdf.cell(50, 8, limpiar(pedido.clienteId.nombre_cliente), 1)
        pdf.cell(35, 8, pedido.fecha_pedido.strftime('%Y-%m-%d'), 1)
        pdf.cell(35, 8, pedido.get_estado_display(), 1)
        pdf.ln()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos.pdf"'
    response.write(bytes(pdf.output(dest='S')))
    return response


@login_required
def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.select_related('clienteId').all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'
    ws.append(['ID', 'Cliente', 'Fecha', 'Estado'])
    for pedido in pedidos:
        ws.append([
            pedido.pedidoId,
            pedido.clienteId.nombre_cliente,
            pedido.fecha_pedido.strftime('%Y-%m-%d %H:%M:%S'),
            pedido.get_estado_display(),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response


