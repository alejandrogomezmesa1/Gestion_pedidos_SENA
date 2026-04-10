from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

# Endpoint de login JWT por AJAX
from django.contrib.auth import authenticate, login as auth_login
@csrf_exempt
def login_jwt_view(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        user = authenticate(request, username=username, password=password)
        serializer = TokenObtainPairSerializer(data=data)
        if user is not None and serializer.is_valid():
            auth_login(request, user)
            return JsonResponse(serializer.validated_data)
        return JsonResponse({'error': 'Credenciales inválidas'}, status=400)
from django.contrib.auth import logout
from .models import Cliente

def logout_view(request):
    logout(request)
    return redirect('/login/')
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.shortcuts import render, redirect
from rest_framework import generics
from .serializers import UserRegisterSerializer

class UserRegisterAPIView(generics.CreateAPIView):
    serializer_class = UserRegisterSerializer

# ── LOGIN SIMPLE ─────────────────────────────────────────────────────────────
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm

def login_view(request):
    if request.user.is_authenticated:
        return redirect('/')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        login(request, user)
        next_url = request.GET.get('next') or '/'
        return redirect(next_url)
    return render(request, 'clientes/login.html', {'form': form})

def registro_view(request):
    if request.user.is_authenticated:
        return redirect('/')
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('/')
    else:
        form = UserCreationForm()
    return render(request, 'clientes/registro.html', {'form': form})

# ── Importaciones necesarias ─────────────────────────────────────────────────
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required      # Protege vistas de funciones
from django.contrib.auth.mixins import LoginRequiredMixin      # Protege vistas de clases
from django.views.generic import ListView                      # Vista genérica para listar
from fpdf import FPDF                                          # Generación de PDF
from openpyxl import Workbook                                  # Generación de Excel
from .models import Cliente
from .forms import ClienteForm

# ── LISTADO ───────────────────────────────────────────────────────────────────
# LoginRequiredMixin redirige al login si el usuario no está autenticado.
# paginate_by=10 divide los resultados en páginas de 10 registros automáticamente.
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'clientes/listar_cliente.html'
    context_object_name = 'clientes'   # nombre de la variable en el template
    paginate_by = 4
    ordering = ['clienteId']


# ── CREAR ─────────────────────────────────────────────────────────────────────
# Si llega un POST procesa el formulario; si llega GET muestra el formulario vacío.
@login_required
def crear_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)   # Llena el form con los datos enviados
        if form.is_valid():
            form.save()                    # Guarda en la base de datos
            return redirect('listar_clientes')
    else:
        form = ClienteForm()               # Formulario vacío para mostrar
    return render(request, 'clientes/crear_cliente.html', {'form': form})


# ── VER DETALLE ───────────────────────────────────────────────────────────────
# get_object_or_404 busca el registro por su pk; si no existe devuelve error 404.
@login_required
def ver_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    return render(request, 'clientes/ver_cliente.html', {'cliente': cliente})


# ── ACTUALIZAR ────────────────────────────────────────────────────────────────
# instance=cliente vincula el formulario al registro existente para editarlo.
@login_required
def actualizar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)  # Actualiza el registro existente
        if form.is_valid():
            form.save()
            return redirect('listar_clientes')
    else:
        form = ClienteForm(instance=cliente)  # Muestra el form con los datos actuales
    return render(request, 'clientes/actualizar_cliente.html', {'form': form, 'cliente': cliente})


# ── ELIMINAR ──────────────────────────────────────────────────────────────────
@login_required
def eliminar_cliente(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    error = None
    if request.method == 'POST':
        try:
            cliente.delete()
            return redirect('listar_clientes')
        except Exception as e:
            error = 'No se puede eliminar el cliente porque tiene pedidos asociados.'
    return render(request, 'clientes/eliminar_cliente.html', {'cliente': cliente, 'error': error})


# ── EXPORTAR PDF ──────────────────────────────────────────────────────────────
# La fuente Arial no soporta tildes en fpdf2, por eso se reemplazan con limpiar().
@login_required
def exportar_clientes_pdf(request):
    def limpiar(texto):
        # Reemplaza caracteres con tilde por su equivalente sin tilde
        reemplazos = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n',
                      'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ñ':'N'}
        for k, v in reemplazos.items():
            texto = texto.replace(k, v)
        return texto

    clientes = Cliente.objects.all()
    pdf = FPDF()
    pdf.add_page()
    # Calcula el ancho útil de la página (total - márgenes)
    ancho = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(ancho, 10, 'Reporte de Clientes', ln=True)
    # Cabecera de tabla
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(15, 8, 'ID', 1)
    pdf.cell(45, 8, 'Nombre', 1)
    pdf.cell(50, 8, 'Email', 1)
    pdf.cell(45, 8, 'Dirección', 1)
    pdf.cell(30, 8, 'Teléfono', 1)
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    for cliente in clientes:
        pdf.cell(15, 8, str(cliente.clienteId), 1)
        pdf.cell(45, 8, limpiar(cliente.nombre_cliente), 1)
        pdf.cell(50, 8, limpiar(cliente.email), 1)
        pdf.cell(45, 8, limpiar(cliente.direccion), 1)
        pdf.cell(30, 8, limpiar(cliente.telefono), 1)
        pdf.ln()

    # Devuelve el PDF como descarga directa al navegador
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'
    response.write(bytes(pdf.output(dest='S')))
    return response


# ── EXPORTAR EXCEL ────────────────────────────────────────────────────────────
@login_required
def exportar_clientes_excel(request):
    clientes = Cliente.objects.all()
    wb = Workbook()          # Crea un libro de Excel en memoria
    ws = wb.active           # Hoja activa del libro
    ws.title = 'Clientes'
    ws.append(['ID', 'Nombre', 'Email', 'Dirección', 'Teléfono'])  # Fila de cabeceras
    for cliente in clientes:
        ws.append([
            cliente.clienteId,
            cliente.nombre_cliente,
            cliente.email,
            cliente.direccion,
            cliente.telefono,
        ])

    # Devuelve el archivo Excel como descarga directa
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response

from rest_framework import generics
from django.contrib.auth.models import User
from .serializers import UserSerializer

class UserListAPIView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer